"""流式导出测试。

保证 openpyxl write_only 流式写正确：行数、表头、宽字段/结构字段、缺失键。
内存特性（常量内存）由人工大规模 smoke 验证，这里固化功能正确性回归。
"""
from openpyxl import load_workbook

from datapulse.pipeline.eval_engine import (
    _ROW_COLUMNS,
    _cell,
    _row_record,
    _stream_to_xlsx,
)


def _fake_row(i: int) -> dict:
    return {
        "session": f"s{i}", "turn": i % 3 + 1, "question": f"问题{i}",
        "j_intent": "资产查询", "dispatch_scene": "正常",
        "j_resolved": "是", "j_resolved_raw": "yes",
        "judge": {"should_dispatch_to_bu": True, "dispatch_reason": "理由",
                  "resolved_reason": "已解决", "needs_human_review": False},
        "dispatched_to_bu": True,
        "gold": {"dispatch": "是", "resolved": "是"},
        "answer_text": "答案原文" * 50,   # 宽字段
    }


def test_stream_writes_all_rows(tmp_path):
    out = tmp_path / "rows.xlsx"
    n = 1000
    _stream_to_xlsx(out, _ROW_COLUMNS, (_row_record(_fake_row(i)) for i in range(n)))

    wb = load_workbook(out, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == n + 1                 # 含表头
    assert list(rows[0]) == _ROW_COLUMNS      # 表头顺序一致
    assert rows[1][0] == "s0"                 # 首行首列
    assert rows[1][_ROW_COLUMNS.index("答案原文")] == "答案原文" * 50
    wb.close()


def test_cell_coerces_non_scalar():
    # xlsx 单元格只接受标量：None 归空，dict/list 转字符串
    assert _cell(None) == ""
    assert _cell("x") == "x"
    assert _cell(True) is True
    assert _cell(3) == 3
    assert isinstance(_cell({"a": 1}), str)
    assert isinstance(_cell([1, 2]), str)


def test_row_record_missing_keys_safe(tmp_path):
    # judge 非 dict（出错行）/ 缺字段时不应抛错，写空串
    bad = {"session": "s", "turn": 1, "question": "q", "j_intent": "",
           "j_resolved": "", "judge": "ERROR-STRING",
           "gold": {}, "answer_text": ""}
    rec = _row_record(bad)
    out = tmp_path / "bad.xlsx"
    _stream_to_xlsx(out, _ROW_COLUMNS, iter([rec]))
    wb = load_workbook(out, read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    assert len(rows) == 2
    wb.close()
